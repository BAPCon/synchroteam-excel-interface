'''
Created on Nov 18, 2021

@author: nifty
'''

from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import json
import traceback
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import sys


import base64
from STConnector import DataPull
app = QApplication([])
window = QMainWindow()
layout = QVBoxLayout()

class job_fields_obj():
    def __init__(self, frame, label_text, field):
        sizePolicy1 = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        sizePolicy1.setHorizontalStretch(0)
        sizePolicy1.setVerticalStretch(0)
        #sizePolicy1.setHeightForWidth(self.scrollAreaWidgetContents.sizePolicy().hasHeightForWidth())
        
        font1 = QFont()
        font1.setPointSize(9)
        font1.setBold(False)
        font1.setWeight(50)
        
        
        self.label_5 = QLabel(frame)
        self.label_5.setObjectName(u"label_"+label_text)
        self.label_5.setGeometry(QRect(10, 12, 191, 16))
        sizePolicy1.setHeightForWidth(self.label_5.sizePolicy().hasHeightForWidth())
        self.label_5.setSizePolicy(sizePolicy1)
        self.label_5.setBaseSize(QSize(15, 10))
        self.label_5.setFont(font1)
        self.label_5.setScaledContents(False)
        self.label_5.setAlignment(Qt.AlignLeading|Qt.AlignLeft|Qt.AlignTop)
        self.label_5.setText(QCoreApplication.translate("MainWindow", field, None))
        self.checkBox_2 = QCheckBox(frame)
        self.checkBox_2.setObjectName(u"checkBox_"+label_text)
        self.checkBox_2.setGeometry(QRect(200, 6, 41, 31))
        print("MADEE")
        
class Ui_MainWindow(object):
    
    def gen_file(self):
        self.selectedFields = []
        for f in self.job_frame_list:
                if f.children()[1].isChecked():
                    self.selectedFields.append(f.children()[0].text())
        print(self.selectedFields)
        self.lim = False
        self.sync_fields()
        wb = Workbook()
        dest_file = self.lineEdit.text()+".xlsx"
        
        report_worksheet = wb.active
        report_worksheet.title = "Report"
        
        data_worksheet = wb.create_sheet(title = "Data")
        
        r = 2
        try:
            for sf in self.selectedFields:
                data_worksheet.cell(row=1, column=self.selectedFields.index(sf)+1, value = sf)
            for job in self.job_records:
                lrr = []
                c = 1
                for sf in self.selectedFields:
                    
                    res = ""
                    try:
                        if "_" in sf:
                            sf_l = sf.split("_")
                            if sf_l[0] == "custom":
                                for l in job['customFieldValues']:
                                    if l['label'] == sf_l[1]:
                                        res = l['value']
                            else:
                                res = job[sf_l[0]][sf_l[1]]
                        else:
                            res = job[sf]
                    except:
                        res = ""
                    data_worksheet.cell(row=r, column=c, value=res)
                    lrr.append(res)
                    c += 1
                print("|".join(lrr)+"\n")
                    
                        
                    
                r += 1
        except Exception as e:
            print(e)
        wb.save(filename = dest_file)
                
        
        
    def load_job_fields(self, fields_list):
        self.label_25.setText("")
        self.record_count.setText(str(self.job_count))
        iter = 0
        try:
            
            for f in self.job_frame_list:
                for c in f.children():
                    c.deleteLater()
                f.deleteLater()
                f = None
            self.job_frame_list = []
            self.fields_list = fields_list + open("fields.txt","r").read().split("\n")
            print(self.fields_list)
            
            
            
            self.scrollAreaWidgetContents.setGeometry(QRect(0, 0, 257, 50*len(self.fields_list)))
            for field in self.fields_list:
                print(field)
                iter += 1
                exec_query = """
self.frame_"""+str(iter)+""" = QFrame(self.scrollAreaWidgetContents)
self.frame_"""+str(iter)+""".setObjectName(u"frame_"""+str(iter)+"""")
self.frame_"""+str(iter)+""".setFrameShape(QFrame.StyledPanel)
self.frame_"""+str(iter)+""".setFrameShadow(QFrame.Raised)
if """+str(iter)+"""%2 == 0:
    self.frame_"""+str(iter)+""".setPalette(self.lightgrey_palette)
    self.frame_"""+str(iter)+""".setAutoFillBackground(True)
else:
    self.frame_"""+str(iter)+""".setPalette(self.white_palette)
    self.frame_"""+str(iter)+""".setAutoFillBackground(True)
self.verticalLayout.addWidget(self.frame_"""+str(iter)+""")
rj = job_fields_obj(self.frame_"""+str(iter)+""", str("""+str(iter)+"""), field)
self.job_frame_list.append(self.frame_"""+str(iter)+""")
#print("AWLLw")
            """
                exec(exec_query)
        except Exception as e:
            print(e)
    def toggle_date_range(self):
        try:
            if self.date_active:
                self.date_active = False
                self.dateEdit_2.setEnabled(False)
                self.dateEdit.setEnabled(False)
            else:
                self.date_active = True
                self.dateEdit_2.setEnabled(True)
                self.dateEdit.setEnabled(True)
        except Exception as e:
            print(e)
    def toggle_max_num(self):
        try:
            if self.max_num_active:
                self.max_num_active = False
                self.spinBox_maxnum.setEnabled(False)
            else:
                self.max_num_active = True
                self.spinBox_maxnum.setEnabled(True)
        except Exception as e:
            print(e)    
  

    def sync_fields(self):
        try:
            
            self.apikey = self.lineEdit_apikey.text()
            self.sitename = self.lineEdit_sitename.text()
            message_bytes = str(self.sitename+":"+self.apikey).encode('ascii')
            self.job_count = 0
            print(message_bytes)
            base64_bytes = base64.urlsafe_b64encode(message_bytes)
            x = 1
            self.label_load = [".","..","..."]
            self.fields_list = []
            self.job_records = []
            while x<100:
                
                self.recent_pull = DataPull.pull_jobs(x, base64_bytes)
                self.job_records += self.recent_pull['data']
                self.job_count += len(self.recent_pull['data'])
                self.label_25.setText(self.label_load[(x-1)%3])
                
                print(self.recent_pull)
                if len(self.recent_pull['data']) == 0:
                    x = 1001
                    continue
                x += 1
                
        except Exception as e:
            print(e)
    def get_customs(self):
        try:
            self.apikey = self.lineEdit_apikey.text()
            self.sitename = self.lineEdit_sitename.text()
            message_bytes = str(self.sitename+":"+self.apikey).encode('ascii')
            
            base64_bytes = base64.urlsafe_b64encode(message_bytes)
            data = DataPull.pull_customs(base64_bytes)
            print(data)
            ld = []
            for l in data['data']:
                ld.append("custom_"+l['label'])
            self.load_job_fields(ld)
        except Exception as e:
            print(e)
    def restart_prog(self):
        os.execl(sys.executable, sys.executable, *sys.argv)
    def open_preset(self):
        try:
            name = QFileDialog.getOpenFileName(caption='Open File')
            print(name)
            file = open(name[0],'r')
            print("Awawr")
            json_data = json.loads(file.read().replace("'",'"').replace("True","true").replace("False","false"))
            
            self.lineEdit_apikey.setText(json_data['apikey'])
            self.lineEdit_sitename.setText(json_data['sitename'])
            self.lineEdit.setText(json_data['workbookname'])
            self.date_active = json_data['date_active']
            self.max_num_active = json_data['limit_amount_active']
            
            
            self.checkBox_maxnum.setChecked(self.max_num_active)
            self.checkBox_4.setChecked(self.date_active)
            self.spinBox_maxnum.setValue(json_data['max_records'])
            
            self.dateEdit.setEnabled(self.date_active)
            self.dateEdit_2.setEnabled(self.date_active)
            
            
                
            print(json_data['date_to'])
            print(QDateTime.fromString(json_data['date_to']))
            
            self.dateEdit.setDateTime(QDateTime.fromString(json_data['date_from']))
            self.dateEdit_2.setDateTime(QDateTime.fromString(json_data['date_to']))
            self.selectedFields = json_data['selected_fields']
            self.fields_list = json_data['fields_list']
            self.job_count = json_data['job_count']
            self.load_job_fields(self.fields_list)
                
            print( self.selectedFields)
            i = 1
            while i <= len(self.fields_list):
                exec("""

print(self.frame_"""+str(i)+""".children()[0].text())
if self.frame_"""+str(i)+""".children()[0].text() in self.selectedFields:
    self.frame_"""+str(i)+""".children()[1].setChecked(True)
                """)
                i += 1
        
            
            file.close()
        except Exception as e:
            print(e)
    def save_preset(self):
        try:
            name = QFileDialog.getSaveFileName(caption='Save File')
            print(name)
            file = open(name[0]+".txt",'w')
            print("Awawr")
            self.selectedFields = []
            for f in self.job_frame_list:
                    if f.children()[1].isChecked():
                        self.selectedFields.append(f.children()[0].text())
            
            text = {
                "apikey": self.lineEdit_apikey.text(),
                "sitename":self.lineEdit_sitename.text(),
                "workbookname": self.lineEdit.text(),
                "date_active": self.date_active,
                "limit_amount_active": self.max_num_active,
                "date_from": str(self.dateEdit.dateTime().toString()),
                "date_to": str(self.dateEdit_2.dateTime().toString()),
                "max_records": self.spinBox_maxnum.value(),
                "selected_fields": self.selectedFields,
                "fields_list": self.fields_list,
                "job_count": self.job_count
                }

            print(str(text))
            file.write(str(text))
            file.close()
        except Exception as e:
            print(e)
    def setupUi(self, MainWindow):
        if not MainWindow.objectName():
            MainWindow.setObjectName(u"MainWindow")
        MainWindow.setEnabled(True)
        MainWindow.resize(802, 622)
        self.date_active = False
        self.max_num_active = False
        self.job_frame_list = []
        self.fields_list = []
        self.actionNew = QAction(MainWindow)
        self.actionNew.setObjectName(u"actionNew")
        self.actionNew.setEnabled(False)
        self.actionNew.triggered.connect(self.restart_prog)
        self.selectedFields = []
        self.actionSave_Preset = QAction(MainWindow)
        self.actionSave_Preset.setObjectName(u"actionSave_Preset")
        self.actionSave_Preset.triggered.connect(self.save_preset)
        self.actionSave_Preset.setShortcut("Ctrl+S")
        self.actionSave_Preset.setStatusTip('Save File')
        self.actionSave_Preset.setEnabled(False)
        self.actionOpen_Preset = QAction(MainWindow)
        self.actionOpen_Preset.setObjectName(u"actionOpen_Preset")
        self.actionOpen_Preset.triggered.connect(self.open_preset)
        self.actionOpen_Preset.setShortcut("Ctrl+O")
        self.actionOpen_Preset.setStatusTip('Open File')
        self.actionOpen_Preset.setEnabled(False)
        self.actionAbout = QAction(MainWindow)
        self.actionAbout.setObjectName(u"actionAbout")
        self.actionAbout.setEnabled(False)
        self.actionExit = QAction(MainWindow)
        self.actionExit.setObjectName(u"actionExit")
        self.centralwidget = QWidget(MainWindow)
        self.centralwidget.setObjectName(u"centralwidget")
        self.centralwidget.setEnabled(True)
        self.actionAbout.setEnabled(False)
        palette = QPalette()
        
        brush = QBrush(QColor(255, 255, 255, 255))
        brush.setStyle(Qt.SolidPattern)
        palette.setBrush(QPalette.Active, QPalette.Base, brush)
        palette.setBrush(QPalette.Active, QPalette.Window, brush)
        palette.setBrush(QPalette.Inactive, QPalette.Base, brush)
        palette.setBrush(QPalette.Inactive, QPalette.Window, brush)
        palette.setBrush(QPalette.Disabled, QPalette.Base, brush)
        palette.setBrush(QPalette.Disabled, QPalette.Window, brush)
        self.white_palette = palette
        self.lightgrey_palette = QPalette()
        
        brush = QBrush(QColor(230, 230, 230, 255))
        brush.setStyle(Qt.SolidPattern)
        self.lightgrey_palette.setBrush(QPalette.Active, QPalette.Base, brush)
        self.lightgrey_palette.setBrush(QPalette.Active, QPalette.Window, brush)
        self.lightgrey_palette.setBrush(QPalette.Inactive, QPalette.Base, brush)
        self.lightgrey_palette.setBrush(QPalette.Inactive, QPalette.Window, brush)
        self.lightgrey_palette.setBrush(QPalette.Disabled, QPalette.Base, brush)
        self.lightgrey_palette.setBrush(QPalette.Disabled, QPalette.Window, brush)
        
        self.centralwidget.setPalette(palette)
        font = QFont()
        font.setBold(False)
        font.setWeight(50)
        self.centralwidget.setFont(font)
        self.centralwidget.setAutoFillBackground(False)
        self.gridLayoutWidget = QWidget(self.centralwidget)
        self.gridLayoutWidget.setObjectName(u"gridLayoutWidget")
        self.gridLayoutWidget.setGeometry(QRect(520, 0, 281, 571))
        self.gridLayout = QGridLayout(self.gridLayoutWidget)
        self.gridLayout.setObjectName(u"gridLayout")
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.frame = QFrame(self.gridLayoutWidget)
        self.frame.setObjectName(u"frame")
        self.frame.setFrameShape(QFrame.StyledPanel)
        self.frame.setFrameShadow(QFrame.Raised)
        self.scrollArea = QScrollArea(self.frame)
        self.scrollArea.setObjectName(u"scrollArea")
        self.scrollArea.setGeometry(QRect(0, 30, 280, 511))
        sizePolicy = QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(2)
        sizePolicy.setVerticalStretch(3)
        sizePolicy.setHeightForWidth(self.scrollArea.sizePolicy().hasHeightForWidth())
        self.scrollArea.setSizePolicy(sizePolicy)
        self.scrollArea.setMinimumSize(QSize(0, 0))
        palette1 = QPalette()
        palette1.setBrush(QPalette.Active, QPalette.Base, brush)
        brush1 = QBrush(QColor(220, 220, 220, 255))
        brush1.setStyle(Qt.SolidPattern)
        palette1.setBrush(QPalette.Active, QPalette.Window, brush1)
        palette1.setBrush(QPalette.Inactive, QPalette.Base, brush)
        palette1.setBrush(QPalette.Inactive, QPalette.Window, brush1)
        palette1.setBrush(QPalette.Disabled, QPalette.Base, brush1)
        palette1.setBrush(QPalette.Disabled, QPalette.Window, brush1)
        self.scrollArea.setPalette(palette1)
        self.scrollArea.setFrameShape(QFrame.StyledPanel)
        self.scrollArea.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scrollArea.setSizeAdjustPolicy(QAbstractScrollArea.AdjustIgnored)
        self.scrollArea.setWidgetResizable(False)
        self.scrollArea.setAlignment(Qt.AlignLeading|Qt.AlignLeft|Qt.AlignTop)
        self.scrollAreaWidgetContents = QWidget()
        self.scrollAreaWidgetContents.setObjectName(u"scrollAreaWidgetContents")
        self.scrollAreaWidgetContents.setEnabled(True)
        self.scrollAreaWidgetContents.setGeometry(QRect(0, 0, 257, 120))
        sizePolicy1 = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        sizePolicy1.setHorizontalStretch(0)
        sizePolicy1.setVerticalStretch(0)
        sizePolicy1.setHeightForWidth(self.scrollAreaWidgetContents.sizePolicy().hasHeightForWidth())
        self.scrollAreaWidgetContents.setSizePolicy(sizePolicy1)
        self.verticalLayout = QVBoxLayout(self.scrollAreaWidgetContents)
        self.verticalLayout.setObjectName(u"verticalLayout")
        
        font1 = QFont()
        font1.setPointSize(9)
        font1.setBold(False)
        font1.setWeight(50)
        
        
        
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        
        self.pushButton = QPushButton(self.frame)
        self.pushButton.setObjectName(u"pushButton")
        self.pushButton.setGeometry(QRect(110, 540, 171, 31))
        
        self.refresh_fields_button = QPushButton(self.centralwidget)
        self.refresh_fields_button.setObjectName(u"refresh_fields_button")
        self.refresh_fields_button.setGeometry(QRect(440, 513, 61, 31))
        
        icon = QIcon()
        
        icon.addFile("Downloads/Chevron/Week 2/1200px-Refresh_icon.svg.png", QSize(), QIcon.Normal, QIcon.Off)
        self.refresh_fields_button.setIcon(icon)
        self.refresh_fields_button.setText("Refresh")
        self.label = QLabel(self.frame)
        self.label.setObjectName(u"label")
        self.label.setGeometry(QRect(10, 0, 91, 31))
        font2 = QFont()
        font2.setPointSize(9)
        font2.setBold(True)
        font2.setWeight(75)
        self.label.setFont(font2)
        self.line = QFrame(self.frame)
        self.line.setObjectName(u"line")
        self.line.setGeometry(QRect(90, 0, 20, 31))
        self.line.setFrameShape(QFrame.VLine)
        self.line.setFrameShadow(QFrame.Sunken)
        

        self.gridLayout.addWidget(self.frame, 1, 0, 1, 1)

        self.label_3 = QLabel(self.centralwidget)
        self.label_3.setObjectName(u"label_3")
        self.label_3.setGeometry(QRect(10, 30, 211, 21))
        font4 = QFont()
        font4.setPointSize(12)
        font4.setBold(True)
        font4.setWeight(75)
        self.label_3.setFont(font4)
        self.label_4 = QLabel(self.centralwidget)
        self.label_4.setObjectName(u"label_4")
        self.label_4.setGeometry(QRect(10, 70, 251, 21))
        self.label_4.setFont(font1)
        self.lineEdit = QLineEdit(self.centralwidget)
        self.lineEdit.setObjectName(u"lineEdit")
        self.lineEdit.setGeometry(QRect(260, 70, 131, 22))
        
        self.label_7 = QLabel(self.centralwidget)
        self.label_7.setObjectName(u"label_7")
        self.label_7.setGeometry(QRect(10, 180, 251, 21))
        self.label_7.setFont(font1)
        
        self.label_maxnum = QLabel(self.centralwidget)
        self.label_maxnum.setObjectName(u"label_maxnum")
        self.label_maxnum.setGeometry(QRect(10, 240, 251, 21))
        self.label_maxnum.setFont(font1)
        self.label_maxnum.setText("# Records")
        
        self.checkBox_maxnum = QCheckBox(self.centralwidget)
        self.checkBox_maxnum.setObjectName(u"checkBox_maxnum")
        self.checkBox_maxnum.setGeometry(QRect(100, 243, 41, 20))
        
        self.spinBox_maxnum = QSpinBox(self.centralwidget)
        self.spinBox_maxnum.setObjectName(u"spinBox_maxnum")
        self.spinBox_maxnum.setGeometry(QRect(180, 240, 151, 22))
        self.spinBox_maxnum.setSingleStep(5)
        self.spinBox_maxnum.setMaximum(10000)
        self.spinBox_maxnum.setEnabled(False)
        
        self.label_8 = QLabel(self.centralwidget)
        self.label_8.setObjectName(u"label_8")
        self.label_8.setGeometry(QRect(10, 140, 211, 21))
        self.label_8.setFont(font4)
        self.line_2 = QFrame(self.centralwidget)
        self.line_2.setObjectName(u"line_2")
        self.line_2.setGeometry(QRect(10, 110, 481, 21))
        self.line_2.setFrameShape(QFrame.HLine)
        self.line_2.setFrameShadow(QFrame.Sunken)
        self.checkBox_4 = QCheckBox(self.centralwidget)
        self.checkBox_4.setObjectName(u"checkBox_4")
        self.checkBox_4.setGeometry(QRect(100, 183, 41, 20))
        self.dateEdit = QDateEdit(self.centralwidget)
        self.dateEdit.setObjectName(u"dateEdit")
        self.dateEdit.setGeometry(QRect(220, 180, 110, 22))
        self.dateEdit.setEnabled(False)
        self.dateEdit_2 = QDateEdit(self.centralwidget)
        self.dateEdit_2.setObjectName(u"dateEdit_2")
        self.dateEdit_2.setGeometry(QRect(380, 180, 110, 22))
        self.dateEdit_2.setEnabled(False)
        self.label_9 = QLabel(self.centralwidget)
        self.label_9.setObjectName(u"label_9")
        self.label_9.setGeometry(QRect(350, 180, 21, 20))
        self.label_10 = QLabel(self.centralwidget)
        self.label_10.setObjectName(u"label_10")
        self.label_10.setGeometry(QRect(180, 180, 31, 20))
        self.label_11 = QLabel(self.centralwidget)
        self.label_11.setObjectName(u"label_11")
        self.label_11.setGeometry(QRect(10, 520, 91, 16))
        font5 = QFont()
        font5.setBold(True)
        font5.setWeight(75)
        self.label_11.setFont(font5)
        self.lineEdit_sitename = QLineEdit(self.centralwidget)
        self.lineEdit_sitename.setObjectName(u"lineEdit_sitename")
        self.lineEdit_sitename.setGeometry(QRect(110, 520, 113, 21))
        self.lineEdit_apikey = QLineEdit(self.centralwidget)
        self.lineEdit_apikey.setObjectName(u"lineEdit_apikey")
        self.lineEdit_apikey.setGeometry(QRect(310, 520, 113, 21))
        self.label_12 = QLabel(self.centralwidget)
        self.label_12.setObjectName(u"label_12")
        self.label_12.setGeometry(QRect(240, 520, 61, 16))
        self.label_12.setFont(font5)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QMenuBar(MainWindow)
        self.menubar.setObjectName(u"menubar")
        self.menubar.setGeometry(QRect(0, 0, 802, 26))
        self.menuww = QMenu(self.menubar)
        self.menuww.setObjectName(u"menuww")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QStatusBar(MainWindow)
        self.statusbar.setObjectName(u"statusbar")
        self.statusbar.setCursor(QCursor(Qt.ArrowCursor))
        self.statusbar.setMouseTracking(False)
        self.statusbar.setAutoFillBackground(False)
        MainWindow.setStatusBar(self.statusbar)

        self.menubar.addAction(self.menuww.menuAction())
        self.menuww.addAction(self.actionNew)
        self.menuww.addAction(self.actionSave_Preset)
        self.menuww.addAction(self.actionOpen_Preset)
        self.menuww.addSeparator()
        self.menuww.addAction(self.actionAbout)
        self.menuww.addAction(self.actionExit)
        
        self.label_25 = QLabel(self.centralwidget)
        self.label_25.setObjectName(u"label_25")
        self.label_25.setGeometry(QRect(440, 490, 61, 20))
        font6 = QFont()
        font6.setPointSize(10)
        font6.setBold(True)
        font6.setWeight(75)
        font6.setKerning(True)
        
        font6.setStyleStrategy(QFont.NoAntialias)
        self.label_25.setFont(font6)
        self.label_25.setLayoutDirection(Qt.LeftToRight)
        self.label_25.setAlignment(Qt.AlignCenter)
        self.label_25.setText("...")
        
        self.record_count = QLabel(self.frame)
        self.record_count.setObjectName(u"record_count")
        self.record_count.setGeometry(QRect(0, 540, 111, 31))
        self.record_count.setPalette(self.white_palette)
        
        self.retranslateUi(MainWindow)
        self.lim = True
        self.refresh_fields_button.clicked.connect(self.get_customs)
        self.pushButton.clicked.connect(self.gen_file)
        self.checkBox_4.clicked.connect(self.toggle_date_range)
        self.checkBox_maxnum.clicked.connect(self.toggle_max_num)
        QMetaObject.connectSlotsByName(MainWindow)
        self.job_count = 20
        self.load_job_fields([])
    # setupUi
    

    def retranslateUi(self, MainWindow):
        MainWindow.setWindowTitle(QCoreApplication.translate("MainWindow", u"MainWindow", None))
        self.actionNew.setText(QCoreApplication.translate("MainWindow", u"New Preset", None))
        self.actionSave_Preset.setText(QCoreApplication.translate("MainWindow", u"Save Preset", None))
        self.actionOpen_Preset.setText(QCoreApplication.translate("MainWindow", u"Open Preset", None))
        self.actionAbout.setText(QCoreApplication.translate("MainWindow", u"About", None))
        self.actionExit.setText(QCoreApplication.translate("MainWindow", u"Exit", None))
        
        #self.lineEdit_apikey.setText("4b61e79f-35fe-4cb0-ac62-2d86b34c49d1")
        #self.lineEdit_sitename.setText("dashboard")
        #self.label_6.setText(QCoreApplication.translate("MainWindow", u"TextLabel", None))
        #self.checkBox_3.setText("")
        #self.label_5.setText(QCoreApplication.translate("MainWindow", u"TextLabel", None))
        #self.checkBox_2.setText("")
        self.pushButton.setText(QCoreApplication.translate("MainWindow", u"Generate Excel Workbook", None))
#if QT_CONFIG(tooltip)
        self.refresh_fields_button.setToolTip(QCoreApplication.translate("MainWindow", u"Refresh Job Fields", None))
#endif // QT_CONFIG(tooltip)
        self.refresh_fields_button.setText("Connect")
        self.label.setText(QCoreApplication.translate("MainWindow", u"Job Fields", None))
        self.label_3.setText(QCoreApplication.translate("MainWindow", u"Workbook Options", None))
        self.label_4.setText(QCoreApplication.translate("MainWindow", u"Workbook Name (exclude extension)", None))
        self.lineEdit.setText("")
        self.label_7.setText(QCoreApplication.translate("MainWindow", u"Date Range", None))
        self.label_8.setText(QCoreApplication.translate("MainWindow", u"Job Filtering", None))
        self.checkBox_4.setText("")
        self.label_9.setText(QCoreApplication.translate("MainWindow", u"To", None))
        self.label_10.setText(QCoreApplication.translate("MainWindow", u"From", None))
        self.label_11.setText(QCoreApplication.translate("MainWindow", u"Website Name", None))
        #self.lineEdit_apikey.setText("")
        self.label_12.setText(QCoreApplication.translate("MainWindow", u"API Key", None))
        self.menuww.setTitle(QCoreApplication.translate("MainWindow", u"File", None))
    # retranslateUi
window_n = Ui_MainWindow()
window_n.setupUi(window)
window.show()
app.exec()
